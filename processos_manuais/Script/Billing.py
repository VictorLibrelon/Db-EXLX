import os
import datetime
import openpyxl
from DBLib import SqlServer

class Billin_:
    
    # Função para preencher um arquivo Excel com os dados do banco
    def preencher_excel(self, file_path, dados):
        # Carregar o arquivo Excel existente
        workbook = openpyxl.load_workbook(file_path)

        # Escolher a planilha dentro do arquivo
        sheet = workbook['Sheet1']

        # Encontrar a primeira célula vazia na coluna A
        for row_num in range(1, sheet.max_row + 1):
            if sheet.cell(row=row_num, column=1).value is None:
                start_row = row_num
                break
        else:
            start_row = sheet.max_row + 1

        # Preencher a tabela com os dados do banco, começando a partir da primeira célula vazia
        for row_data in dados:
            row_data_list = [str(value) for value in row_data]
            for col_num, value in enumerate(row_data_list, start=1):
                sheet.cell(row=start_row, column=col_num, value=value)
            start_row += 1

        # Salvar as alterações no arquivo
        output_path = os.path.join('C:\\Automation\\processos_manuais\\Billing_', datetime.datetime.now().strftime("%Y-%m") + '_' + os.path.basename(file_path))
        workbook.save(output_path)

    def Run(self):
        # Criar uma instância do SqlServer
        sql = SqlServer()

        sql.billing_update_employee_id_and_alias()

        print('#############################################')
        print('Update feito com sucesso ')

        # Especificar o caminho da pasta onde estão os arquivos
        pasta = 'C:\\Automation\\processos_manuais\\Billing_Files_Empty'

        # Listar todos os arquivos na pasta
        arquivos = os.listdir(pasta)

        # Lista com os nomes das funções
        funcoes_sql = [
            "billing_get_aip_users",
            "billing_get_atp_users",
            "billing_get_azure_ad_users",
            "billing_get_delv_users", 
            "billing_get_exchange_users",
            "billing_get_forms_users",
            "billing_get_onde_drive_activity",
            "billing_get_planner_users",
            "billing_get_power_apps_users",
            "billing_get_power_automate_userss",
            "billing_get_powerbi_users",
            "billing_get_share_point_activity",
            "billing_get_stream_users",
            "billing_get_sway_users",
            "billing_get_teams_activity",
            "billing_get_to_do_users",
            "billing_get_yammer_activity"
        ]

        rdpCount = 0

        # Iterar sobre os arquivos
        for arquivo in arquivos:
            caminho_completo = os.path.join(pasta, arquivo)
            
            # Verificar se é um arquivo Excel
            if arquivo.endswith('.xlsx'):
                # Obter o nome da função do procedimento SQL atual
                nome_funcao = funcoes_sql[rdpCount]
                
                # Usar a função getattr para chamar dinamicamente a função
                funcao = getattr(sql, nome_funcao)
                dados_do_banco = funcao()
                
                # Preencher o arquivo Excel com os dados do banco
                self.preencher_excel(caminho_completo, dados_do_banco)
                
                # Incrementar o contador de procedimentos SQL
                rdpCount += 1
